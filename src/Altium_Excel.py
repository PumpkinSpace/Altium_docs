#!/usr/bin/env python
###########################################################################
#(C) Copyright Pumpkin, Inc. All Rights Reserved.
#
#This file may be distributed under the terms of the License
#Agreement provided with this software.
#
#THIS FILE IS PROVIDED AS IS WITH NO WARRANTY OF ANY KIND,
#INCLUDING THE WARRANTY OF DESIGN, MERCHANTABILITY AND
#FITNESS FOR A PARTICULAR PURPOSE.
###########################################################################
"""
@package Altium_Excel.py

Package that manages excel files in the Altium Documentation Module.
"""

__author__ = 'David Wright (david@asteriaec.com)'
__version__ = '0.2.0' #Versioning: http://www.python.org/dev/peps/pep-0386/


#
# -------
# Imports

import xlrd
import os
import sys
sys.path.insert(1, 'src\\')
import shutil
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
import Altium_helpers
import Altium_Files
import Altium_GS

#
# -------
# Constants


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

no_border = Border(left=Side(style=None), 
                   right=Side(style=None), 
                   top=Side(style=None), 
                   bottom=Side(style=None))

# define BOM layaout constants

BOM_cols =  {'Designator': 0,
             'DNP': 1,
             'Description' : 2,
             'Quantity': 3,
             'Manufacturer': 4,
             'Manufacturer_pn': 5,
             'Supplier': 6,
             'Supplier_pn': 7,
             'Sub_Manufacturer': 8, 
             'Sub_Manufacturer_pn' : 9,
             'Sub_Supplier' : 10,
             'Sub_Supplier_pn': 11,
             'Subtotal': 12}

bom_d_col = 0
bom_dnp_col = 1
bom_pn_col = 5
bom_header_rows = 6
bom_comment_col = 2

is_test = False

#
# ----------------
# Public Functions 

def set_directory(directory):
    """
    Function to set the initial directory of the executing code

    @param[in]    directory:  The directory to set.
    @attribute    path:       The path that has been set
    """    
    set_directory.path = directory
# end def

# set the initial value
set_directory.path = os.getcwd()


def log_error(get = False):
    """
    Function to log errors within this module.

    @param[in]    get:        True  = return no_errors without logging an error
                              False = log an error and return nothing (bool)
    @attribute    no_errors:  Whether there have been no errors logged
    @return       (bool)      True  = no errors have been logged.
                              False = Errors have been logged.
    """  
    
    # determine which action to take
    if get:
        # return the state
        return log_error.no_errors
    
    else:
        # log an error
        log_error.no_errors = False
    # end if
# end def

# set the initial value
log_error.no_errors = True


def log_warning(get = False):
    """
    Function to log warnings within this module.

    @param[in] get:          True  = return no_warnings without logging a warning
                             False = log a warning and return nothing (bool)
    @attribute no_warnings:  Whether there have been no errors logged
    @return    (bool)        True  = no errors have been logged.
                             False = Errors have been logged.
    """    
    
    # determine which action to take
    if get:
        # return the state
        return log_warning.no_warnings
    
    else:
        # log a warning
        log_warning.no_warnings = False
    # end if
# end def

# set the initial value
log_warning.no_warnings = True


def set_assy_options(starting_dir, list_0, list_1):
    """
    Function to set the assembly rev options in the ASSY Config document based on 
    what was read from the schematic pages.

    @param[in]   starting_dir:     The Altium project directory (full path) 
                                   (string).
    @param[in]   list_0            List of options corresponding to binary 0 
                                   selections (list of stings).
    @param[in]   list_0            List of options corresponding to binary 0 
                                   selections (list of stings).
    @return      (bool)            True if this function is successful.
    """      
    
    # open the assy config document
    [assy_filename, assy_doc, assy_sheet] = open_assy_config(starting_dir, 'Options')    
    
    # check that it got what it wanted
    if assy_sheet == None:
        return False
    # end if
    
    # empty the options sheet to restore it to it's initial state
    for j in range(2,Altium_GS.max_assy_rev+1):
        assy_sheet.cell(j,2).value = ' '
        assy_sheet.cell(j,3).value = 'No corresponding Assembly Configuration'
    # end for
    
    # fill the sheet with the options extracted from the pdf
    for i in range(0,len(list_0)):
        assy_sheet.cell(i+2,2).value = list_0[i]
        assy_sheet.cell(i+2,3).value = list_1[i]
    # end for
    
    # save the file
    assy_doc.active = 0
    assy_doc.save(starting_dir + '\\' + assy_filename)
    
    # close the doc
    assy_doc.close()
    
    # return successful
    return True
# end def
    
    
def construct_assembly_doc(starting_dir, gerber_dir, output_pdf_dir, part_number):
    """
    Function to build the BOM page of the ASSY Config doc based on the BOMs 
    exported from Altium.

    @param[in]   starting_dir:     The Altium project directory (full path) 
                                   (string).
    @param[in]   gerber_dir:       The location of the BOM documents (full path) 
                                   (string).                             
    @param[in]   output_pdf_dir:   The Location of the pdf files (full path)
                                   (string).
    @param[in]   part_number:      The part number for the design
                                   (string).
    @return      (datetime)        The modified date of the BOM
    @return      (datetime)        The modified date of the DNP bom
    """       
    
    
    # if this failed log an error
    if not log_error(get=True):
        return None, None
    # end if

    # initialise BOM column arrays
    bom_d_list = []
    bom_pn_list = []
    dnp_d_list = []
    dnp_pn_list = []
    
    # fill the BOM column arrays
    [bom_doc, bom_date] = get_bom_lists(gerber_dir, bom_d_list, bom_pn_list)
    [dnp_doc, dnp_date] = get_bom_lists(gerber_dir, dnp_d_list, dnp_pn_list, 
                                        DNP=True)
    
    # if this is a test print the lists
    if is_test:
        for i in bom_d_list: print i
        print '\n'
        for i in dnp_d_list: print i
        print '\n'
    # end if
    
    # if getting the BOM lists threw an error then exit
    if not log_error(get=True):
        return None, None
    # end if    

    # initialise the DNP component list
    comp_dnp_list = []
    
    # iternate through the part numbers in the total list
    for index in range(0,len(dnp_pn_list)):
        if dnp_pn_list[index] not in bom_pn_list:
            # if the part number is not present in the list of only placed
            # then move all such parts to the DNP list
            comp_dnp_list.append(dnp_d_list[index])
            
            # empty the placed list
            dnp_d_list[index] = []
            
        else:
            # the part number is present in the placed list so check every designator
            designator_list = []
            
            # find the designator list in the other BOM
            bom_index = bom_pn_list.index(dnp_pn_list[index])
            
            # check each designator
            for designator in dnp_d_list[index]:
                # compare to the other list
                if designator not in bom_d_list[bom_index]:
                    # it is not found so add the componant to the dnp list
                    designator_list.append(designator)
                    
                    # remove the designator from this list
                    dnp_d_list[index].remove(designator)
                # end if
            # end for
            
            # add to the dnp list
            comp_dnp_list.append(designator_list)
        # end if 
    # end for
    
    # if this is a test then print the lists
    if is_test:
        for i in comp_dnp_list: print i
        print '\n'
        for i in dnp_d_list: print i
        print '\n'        
    # end if
    
    # fill the assy config document with these lists.
    fill_assy_bom(starting_dir, output_pdf_dir, part_number, dnp_d_list, comp_dnp_list, dnp_doc)
    
    # extract all information from the ASSY Config document
    assy_data = extract_assy_config(starting_dir)
    
    # use that data to fill the online BOM
    if Altium_GS.populate_online_bom(set_directory.path, 
                                     get_assembly_number('PART'),
                                     get_assembly_number('ASSY'),
                                     get_assembly_number('REV'),
                                     assy_data) == None:
        log_error()
    #end if
        
    
    # return the modified dates
    return bom_date, dnp_date
#end def


def copy_assy_config(starting_dir):
    """
    Function to open copy the master ASSY Config document to the root folder.

    @param[in]   starting_dir:     The Altium project directory (full path) 
                                   (string).
    """ 
    
    # if the ASSY Config document already exists then delete it
    for filename in os.listdir(starting_dir):
        if ('ASSY' in filename) and (('Config' in filename) or ('REV' in filename)):
            os.remove(starting_dir + '\\' + filename)
        # end if
    # end for    
    
    try:
        shutil.copyfile(set_directory.path + '\\src\\ASSY Config.xlsx',
                        starting_dir + '\\ASSY Config.xlsx')
    
    except:
        print '*** Error: could not copy master ASSY Config document ***'
        log_error()
    # end try
# end def


def extract_assy_config(starting_dir):
    """
    Function to extract all of the salient information from the Assembly 
    Revision Document.

    @param[in]   starting_dir:                The Altium project directory 
                                              (full path) (string).
    @return      (Altium_GS.assembly_info)    All of the extracted information.
    """     
    
    # define container to return
    output_data = Altium_GS.assembly_info()
    
    # open the ASSY config document and extract the BOM sheet
    [assy_filename, assy_doc, bom_sheet] = open_assy_config(starting_dir, 'BOM')    
    
    # check that what was requested was returned
    if bom_sheet == None:
        log_error()
        return None
    # end if
    
    # extract all information from the BOM
    for i in range(bom_header_rows+1,bom_sheet.max_row+1):
        if bom_sheet.cell(i,BOM_cols['Designator']+1).value != None:
            output_data.designators.append(bom_sheet.cell(i,BOM_cols['Designator']+1).value.split(', '))
            
        else:
            output_data.designators.append([])
        # end if
        
        if bom_sheet.cell(i,BOM_cols['DNP']+1).value != None:
            output_data.dnp_designators.append(bom_sheet.cell(i,BOM_cols['DNP']+1).value.split(', '))
            
        else:
            output_data.dnp_designators.append([])
        # end if
        
        output_data.descriptions.append(bom_sheet.cell(i,BOM_cols['Description']+1).value)
        output_data.quantities.append(bom_sheet.cell(i,BOM_cols['Quantity']+1).value)
        output_data.manufacturers.append(bom_sheet.cell(i,BOM_cols['Manufacturer']+1).value)
        output_data.sub_manufacturer.append(bom_sheet.cell(i,BOM_cols['Sub_Manufacturer']+1).value)
        output_data.manufacturer_pns.append(bom_sheet.cell(i,BOM_cols['Manufacturer_pn']+1).value)
        output_data.sub_manufacturer_pns.append(bom_sheet.cell(i,BOM_cols['Sub_Manufacturer_pn']+1).value)
        output_data.sub_supplier_pns.append(bom_sheet.cell(i,BOM_cols['Sub_Supplier_pn']+1).value)
        output_data.sub_suppliers.append(bom_sheet.cell(i,BOM_cols['Sub_Supplier']+1).value)
        output_data.suppliers.append(bom_sheet.cell(i,BOM_cols['Supplier']+1).value)
        output_data.supplier_pns.append(bom_sheet.cell(i,BOM_cols['Supplier_pn']+1).value)
        output_data.subtotals.append(bom_sheet.cell(i,BOM_cols['Subtotal']+1).value)
    # end for
    
    # close the file
    assy_doc.close()    
    
    # open the ASSY Config document and extract the BOM sheet
    [assy_filename, assy_doc, option_sheet] = open_assy_config(starting_dir, 'Options')   
    
    # check that what was requested was returned
    if option_sheet == None:
        log_error()
        return None
    # end if    
    
    # extract all of the Assembly Rev information
    for i in range(2,option_sheet.max_row+1):
        output_data.list_0.append(option_sheet.cell(i,2).value)
        output_data.list_1.append(option_sheet.cell(i,3).value)
    # end for
    
    # close the file
    assy_doc.close() 
    
    return output_data
# end def


#
# ----------------
# Private Functions 

def set_assembly_number(doc):
    """
    Function to prompt the user for the assembly number and then store it.

    @param[in]   starting_dir:     The Altium project directory (full path) 
                                   (string).
    """       
    
    # get the assembly number
    cell_string = repr(doc.cell(1,4)).split('\'')[1]
    
    if cell_string.startswith('710-'):
        set_assembly_number.assy_number = cell_string.split('-')[1]
        
    elif cell_string.startswith('711-'):
        set_assembly_number.assy_number = cell_string.split('-')[1]        
        
    else:
        print "*** Error: no assembly number found in BOM Doc ***"
        log_error()
    # end if
    
    # get part number
    cell_string = repr(doc.cell(0,4)).split('\'')[1]
    
    if cell_string.startswith('705-'):
        set_assembly_number.part_number = cell_string.split('-')[1]
        
    else:
        print "*** Error: no part number found in BOM Doc ***"
        log_error()
    # end if   
    
    # get revision
    cell_string = repr(doc.cell(0,6)).split('\'')[1]
    
    if (cell_string[0].isalpha() and (len(cell_string) == 1)):
        # single character rev
        set_assembly_number.revision = (cell_string + '0')
        
    elif (cell_string[0].isalpha() and cell_string[1].isdigit()):
        # character and number rev
        set_assembly_number.revision = cell_string
        
    else:
        print "*** Error: no revision found in BOM Doc ***"
        log_error()
    # end if    
# end def

# set the initial value
set_assembly_number.assy_number = None
set_assembly_number.part_number = None
set_assembly_number.revision = None

def get_assembly_number(specific_number = 'ASSY'):
    """
    Function to prompt the user for the assembly number and then store it.
    """       
    
    if (specific_number == 'ASSY'):
        if (set_assembly_number.assy_number == None):
            print "*** Error: No Assembly Number has been set ***"
            log_error()
            
        else:
            return set_assembly_number.assy_number
        # end if 
        
    elif (specific_number == 'PART'):
        if (set_assembly_number.part_number == None):
            print "*** Error: No Part Number has been set ***"
            log_error()
            
        else:
            return set_assembly_number.part_number
        # end if    
        
    elif (specific_number == 'REV'):
        if (set_assembly_number.revision == None):
            print "*** Error: No Revision has been set ***"
            log_error()
            
        else:
            return set_assembly_number.revision
        # end if  
        
    else:
        print "*** Error: Invalid input to function ***"
        log_error()
    # end if
# end def


def get_bom_lists(gerber_dir, d_list, pn_list, DNP = False):
    """
    Function to extract the designator and part number lists from a BOM.

    @param[in]   gerber_dir:       The Altium project directory (full path) 
                                   (string).
    @param[out]  d_list:           The list to add the found designator lists to
                                   (list of lists of strings).
    @param[out]  pn_list:          The list to add the found part number lists to
                                   (list of lists of strings).
    @param[in]   DNP:              True  = look in the DNP BOM
                                   False = look in the regular BOM
                                   (bool)
    @return      (worksheet)       The BOM sheet that was opened.
    @return      (mod_date)        The modification date of the BOM.
    """      
    
    filename = ''
    
    # find the BOM docs.
    for name in os.listdir(gerber_dir):
        if (DNP == True and name.endswith('.xls') and not name.endswith(').xls')):
            filename = name
            break
            
        elif (DNP == False and ('Placed Components Only' in name)):
            filename = name
            break
        # end if
    # end for  
    
    # no file was found so log the appropraite error
    if filename == '':
        if DNP:
            print '***  Error: no DNP BOM found ***'
            log_error()
            
        else:
            print '***  Error: no BOM found ***'
            log_error()   
        # end if
        return None, None
    # end if
    
    if is_test:
        print filename
    # end if
    
    try:
        # get the BOM date
        date = Altium_helpers.mod_date(os.path.getmtime(gerber_dir + '\\' + filename),
                                       filename)
        
        # open the BOM sheet
        doc = xlrd.open_workbook(gerber_dir + '\\' + filename).sheet_by_index(0)
    
    except:
        print '***  Error: could not open .xls file ***'
        log_error()    
        return None, None
    # end try
    
    # extract the required information
    for row in range(bom_header_rows, doc.nrows):
        # find the part number in the BOM Doc.
        pn_list.append(extract_items(doc.cell(row,BOM_cols['Manufacturer_pn'])))
        d_list.append(extract_items(doc.cell(row,BOM_cols['Designator'])))
    # end for   
    
    if (DNP == True):
        # extract the assembly number from the BOM
        set_assembly_number(doc)
    # end if
        
    
    # return information
    return doc, date
# end def


def extract_items(cell):
    """
    Function to extract text from a worksheet cell.

    @param[in]   cell:              The cell to get the data from (cell).
    @return      (list of strings)  The extracted information.
    """      
    # extract the text
    cell_string = repr(cell).strip('text:u\'')
    
    # split the list into sub parts
    cell_list = cell_string.split(', ')
    
    # return the list
    return cell_list
# end def


def fill_assy_bom(starting_dir, output_pdf_dir, part_number, dnp_d_list, comp_dnp_list, dnp_doc):
    """
    Function to populate the ASSY Config document with the extracted BOM 
    information.

    @param[in]   starting_dir:     The Altium project directory (full path) 
                                   (string).
    @param[in]   output_ pdf_dir:  The Location to place the pdf files (full path)
                                   (string).
    @param[in]   part_number:      The part number for the design
                                   (string).
    @param[in]   dnp_d_list:       The list of components to place
                                   (list of lists of strings).
    @param[in]   comp_dnp_list:    The list of components not to place
                                   (list of lists of strings).
    @param[in]   dnp_doc:          The BOM that provides all the other 
                                   information (worksheet)
    """ 
    
    # open the ASSY Config document and extract the BOM sheet
    [assy_filename, assy_doc, bom_sheet] = open_assy_config(starting_dir, 'BOM')    
    
    # check that what was requested was returned
    if bom_sheet == None:
        log_error()
        return None
    # end if
    
    # empty bom and remove borders to reset it to empty state
    for i in range(1,bom_sheet.max_row+1):
        for j in range(1,bom_sheet.max_column+1):
            # empty the cell
            bom_sheet.cell(i,j).value = ''
            
            # remove borders
            if i > bom_header_rows:
                bom_sheet.cell(i,j).border = no_border
            #end if
        # end for
    # end for
    
    # full replace all cells in the BOM sheet
    for i in range(0,dnp_doc.nrows):
        for j in range(0,dnp_doc.ncols):
            if (i >= bom_header_rows):
                # into BOM information
                if (j == BOM_cols['Designator']) and (i < dnp_doc.nrows):
                    # add the designators to place
                    bom_sheet.cell(i+1,j+1).value = ', '.join(dnp_d_list[i-6])
                    
                elif ((j == BOM_cols['DNP']) and (i < dnp_doc.nrows-1)):
                    # add the designators not to place
                    bom_sheet.cell(i+1,j+1).value = ', '.join(comp_dnp_list[i-6])   
                    
                elif ((j == BOM_cols['Description']) and (i < dnp_doc.nrows-1)):
                    # tidy up comments
                    bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)   
                    
                    comment_list = bom_sheet.cell(i+1,j+1).value.split(',')
                    
                    if (len(comment_list) > 1):
                        # there is more than one type of comment for this part
                        # so pick one
                        bom_sheet.cell(i+1,j+1).value = comment_list[-1].strip()
                    # end if
                    
                else:
                    # copy the part info from the BOM
                    bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)
                # end if
                
                # add borders to all cells
                bom_sheet.cell(i+1,j+1).border = thin_border
            
            else:
                # copy the BOM info from the BOM
                bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)            
            # end if
        # end for
    # end for
    
    # save the file
    assy_doc.active = 0
    assy_doc.save(starting_dir + '\\' + assy_filename)
    
    # extract just the BOM
    for sheet in assy_doc.sheetnames:
        if sheet != 'BOM':
            sheet_doc = assy_doc[sheet]
            assy_doc.remove(sheet_doc)
        # end if
    # end for
    
    # save the file as just the BOM
    assy_doc.save(output_pdf_dir + '//' + part_number + ' digikey order.xlsx')
        
    # close the file
    assy_doc.close()
# end def


def open_assy_config(starting_dir, sheet = 'BOM'):
    """
    Function to open the ASSY Config document and return the desired sheet.

    @param[in]   starting_dir:     The Altium project directory (full path) 
                                   (string).
    @param[in]   sheet:            The name of the desired sheet in the document
                                   (string).
    @return      (string)          The filename of the document.
    @return      (workbook)        The workbook that was opened.
    @return      (worksheet)       The requested sheet in the document.
    """     
    # find assy config document
    assy_filename = ''
    
    for filename in os.listdir(starting_dir):
        if ('ASSY' in filename) and ('Config' in filename):
            assy_filename = filename
        # end if
    # end for
    
    if assy_filename == '':
        print '***  Error: no ASSY Config doc found ***'
        
        return None, None, None
    # end if    
        
    try:
        # open the assy config document
        assy_doc = openpyxl.load_workbook(starting_dir + '\\' + assy_filename)
        
    except:
        print '***  Error: ASSY Config doc could not be opened ***'
    
        return None, None, None      
    # end try
    
    try:
        # open the BOM sheet
        bom_sheet = assy_doc[sheet]
        
    except:
        print '***  Error: ASSY Config doc is invlaid ***'
    
        return None, None, None       
    # end try
    
    # return all desired information
    return assy_filename, assy_doc, bom_sheet
# end def


def test():
    """
    Test code for this module.
    """
    Altium_helpers.clear_output('\\'.join(os.getcwd().split('\\')[:-1]) + '\\test folder (01234A)', True)
    construct_assembly_doc('\\'.join(os.getcwd().split('\\')[:-1]) + '\\test folder (01234A)')
    
    if not log_error(get=True):
        print '*** ERRORS OCCURRED***'
#end def

if __name__ == '__main__':
    # if this code is not running as an imported module run test code
    is_test = True
    test()
# end if



                


                
        
        
