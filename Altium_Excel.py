import xlrd
import os
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
import Altium_helpers

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

no_border = Border(left=Side(style=None), 
                   right=Side(style=None), 
                   top=Side(style=None), 
                   bottom=Side(style=None))

# define BOM layaout constants
bom_d_col = 0
bom_pn_col = 5
bom_dnp_col = 1
bom_header_rows = 6
bom_comment_col = 2

is_test = False

def log_error(get = False):
    if get:
        return log_error.no_errors
    
    else:
        log_error.no_errors = False
    # end if
# end def
log_error.no_errors = True

def log_warning(get = False):
    if get:
        return log_warning.no_warnings
    
    else:
        log_warning.no_warnings = False
    # end if
# end def
log_warning.no_warnings = True



def extract_items(cell):
    cell_string = repr(cell).strip('text:u\'')
    cell_list = cell_string.split(', ')
    return cell_list
# end def


def get_bom_lists(starting_dir, d_list, pn_list, DNP = False):
    
    # find output directory
    output_dir = Altium_helpers.get_output_dir(starting_dir)
    
    filename = ''
    
    # find two BOM docs.
    for name in os.listdir(output_dir):
        if DNP == True and name.startswith('DNP'):
            filename = name
            
        elif DNP == False and name.endswith('.xls') and not name.startswith('DNP'):
            filename = name
        # end if
    # end for  
    
    if filename == '':
        if DNP:
            print '***  Error: no DNP BOM found ***'
            log_error()
            
        else:
            print '***  Error: no BOM found ***'
            log_error()   
        # end if
        return None, None
    # end if
    
    if is_test:
        print filename
    # end if
    
    try:
        
        # get the BOM date
        date = os.path.getmtime(output_dir + '\\' + filename)
        
        # open the BOM sheet
        doc = xlrd.open_workbook(output_dir + '\\' + filename).sheet_by_index(0)
    
    except:
        print '***  Error: could not open .xls file ***'
        log_error()    
        return None, None
    # end try
    
    for row in range(bom_header_rows, doc.nrows):
        # find the part number in the BOM Doc.
        pn_list.append(extract_items(doc.cell(row,bom_pn_col)))
        d_list.append(extract_items(doc.cell(row,bom_d_col)))
    # end for    
    
    return doc, date
# end def

def open_assy_rev(starting_dir, sheet = 'BOM'):
    # find assy_rev document
    assy_filename = ''
    
    for filename in os.listdir(starting_dir):
        if ('ASSY' in filename) and ('REV' in filename):
            assy_filename = filename
        # end if
    # end for
    
    if assy_filename == '':
        print '***  Error: no ASSY_REV doc found ***'
        
        return None, None, None
    # end if    
        
    try:
        # open the assy_rev document
        assy_doc = openpyxl.load_workbook(starting_dir + '\\' + assy_filename)
        
    except:
        print '***  Error: ASSY_REV doc could not be opened ***'
    
        return None, None, None      
    # end try
    
    try:
        # open the BOM sheet
        bom_sheet = assy_doc[sheet]
        
    except:
        print '***  Error: ASSY_REV doc is invlaid ***'
    
        return None, None, None       
    # end try
    
    return assy_filename, assy_doc, bom_sheet
# end def

def fill_assy_bom(starting_dir, dnp_d_list, comp_dnp_list, dnp_doc):
    
    [assy_filename, assy_doc, bom_sheet] = open_assy_rev(starting_dir, 'BOM')    
    
    if bom_sheet == None:
        log_error()
        return None
    # end if
    
    # empty bom and remove borders to reset it to empty state
    for i in range(1,bom_sheet.max_row+1):
        for j in range(1,bom_sheet.max_column+1):
            bom_sheet.cell(i,j).value = ''
            
            if i > bom_header_rows:
                bom_sheet.cell(i,j).border = no_border
            #end if
        # end for
    # end for
    
    # full replace all cells in the BOM sheet
    for i in range(0,dnp_doc.nrows):
        for j in range(0,dnp_doc.ncols):
            if (i >= bom_header_rows):
                # into BOM information
                if (j == bom_d_col) and (i < dnp_doc.nrows):
                    # add the designators to place
                    bom_sheet.cell(i+1,j+1).value = ', '.join(dnp_d_list[i-6])
                    
                elif ((j == bom_dnp_col) and (i < dnp_doc.nrows-1)):
                    # add the designators not to place
                    bom_sheet.cell(i+1,j+1).value = ', '.join(comp_dnp_list[i-6])   
                    
                elif ((j == bom_comment_col) and (i < dnp_doc.nrows-1)):
                    # tidy up comments
                    bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)   
                    
                    comment_list = bom_sheet.cell(i+1,j+1).value.split(',')
                    
                    if (len(comment_list) > 1):
                        # there is more than one type of comment for this part
                        # so pick one
                        bom_sheet.cell(i+1,j+1).value = comment_list[-1].strip()
                    # end if
                    
                else:
                    # copy the part info from the BOM
                    bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)
                # end if
                
                # add borders to all cells
                bom_sheet.cell(i+1,j+1).border = thin_border
            
            else:
                # copy the BOM info from the BOM
                bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)            
            # end if
        # end for
    # end for
    
    # save the file
    assy_doc.active = 0
    assy_doc.save(starting_dir + '\\' + assy_filename)
    
    assy_doc.close()

# end def

def set_assy_options(starting_dir, list_0, list_1):
    
    [assy_filename, assy_doc, assy_sheet] = open_assy_rev(starting_dir, 'Options')    
    
    if assy_sheet == None:
        return False
    # end if
    
    # empty the options sheet to restore it to it's initial state
    for j in range(2,assy_sheet.max_row+1):
        assy_sheet.cell(j,2).value = ' '
        assy_sheet.cell(j,3).value = 'No corresponding Assembly Revision'
    # end for
    
    # fill the sheet with the options extracted from the pdf
    for i in range(0,len(list_0)):
        assy_sheet.cell(i+2,2).value = list_0[i]
        assy_sheet.cell(i+2,3).value = list_1[i]
    # end for
    
    # save the file
    assy_doc.active = 0
    assy_doc.save(starting_dir + '\\' + assy_filename)
    
    assy_doc.close()
    
    return True
# end def
    
def construct_assembly_doc(starting_dir):
    
    # find output directory
    output_dir = Altium_helpers.get_output_dir(starting_dir)
    
    if not log_error(get=True):
        return None
    # end if

    # initialise BOM column arrays
    bom_d_list = []
    bom_pn_list = []
    dnp_d_list = []
    dnp_pn_list = []
    
    # fill the BOM column arrays
    [bom_doc, bom_date] = get_bom_lists(starting_dir, bom_d_list, bom_pn_list)
    [dnp_doc, dnp_date] = get_bom_lists(starting_dir, dnp_d_list, dnp_pn_list, DNP=True)
    
    if is_test:
        for i in bom_d_list: print i
        print '\n'
        for i in dnp_d_list: print i
        print '\n'
    # end if
    
    if not log_error(get=True):
        return None
    # end if    

    # initialise the DNP component list
    comp_dnp_list = []
    
    # iternate through the part numbers in the total list
    for index in range(0,len(dnp_pn_list)):
        if dnp_pn_list[index] not in bom_pn_list:
            # if the part number is not present in the list of only placed
            # then move all such parts to the DNP list
            comp_dnp_list.append(dnp_d_list[index])
            
            # empty the placed list
            dnp_d_list[index] = []
            
        else:
            # the part number is present in the placed list so check every designator
            designator_list = []
            
            # find the designator list in the other BOM
            bom_index = bom_pn_list.index(dnp_pn_list[index])
            
            # check each designator
            for designator in dnp_d_list[index]:
                # compare to the other list
                if designator not in bom_d_list[bom_index]:
                    # it is not found so add the componant to the dnp list
                    designator_list.append(designator)
                    
                    # remove the designator from this list
                    dnp_d_list[index].remove(designator)
                # end if
            # end for
            
            # add to the dnp list
            comp_dnp_list.append(designator_list)
            
        # end if 
    # end for
    
    if is_test:
        for i in comp_dnp_list: print i
        print '\n'
        for i in dnp_d_list: print i
        print '\n'        
    # end if
    
    fill_assy_bom(starting_dir, dnp_d_list, comp_dnp_list, dnp_doc)
    
    return bom_date, dnp_date
#end def

def test():
    """
    Test code for this module.
    """
    Altium_helpers.clear_output(os.getcwd() + '\\test folder')
    construct_assembly_doc(os.getcwd() + '\\test folder')
    
    if not log_error(get=True):
        print '*** ERRORS OCCURRED***'
#end def

if __name__ == '__main__':
    # if this code is not running as an imported module run test code
    is_test = True
    test()
# end if



                


                
        
        
