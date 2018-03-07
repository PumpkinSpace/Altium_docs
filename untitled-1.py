import xlrd
import os
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

output_dir = 'C:\Users\Asteria\Dropbox\Satellite\Pumpkin PCBs\ADCS Interface Module 2 (01845B)\Project Outputs'

dnp_filename = ''
bom_filename = ''

# find two BOM docs.
for filename in os.listdir(output_dir):
    if filename.startswith('DNP'):
        dnp_filename = filename
        
    elif filename.endswith('.xls'):
        bom_filename = filename
    #end if
#end for

# open the two BOM docs
dnp_doc = xlrd.open_workbook(output_dir + '\\' + dnp_filename).sheet_by_index(0)
bom_doc = xlrd.open_workbook(output_dir + '\\' + bom_filename).sheet_by_index(0)

bom_d_list = []
bom_pn_list = []
dnp_d_list = []
dnp_pn_list = []

def extract_items(cell):
    cell_string = repr(cell).strip('text:u\'')
    cell_list = cell_string.split(', ')
    return cell_list
# end def

for bom_row in range(6, bom_doc.nrows):
    # find the part number in the BOM Doc.
    bom_pn_list = bom_pn_list + [extract_items(bom_doc.cell(bom_row,5))]
    bom_d_list = bom_d_list + [extract_items(bom_doc.cell(bom_row,0))]
#end for

for dnp_row in range(6, dnp_doc.nrows):
    # find the part number in the BOM Doc.
    dnp_pn_list = dnp_pn_list + [extract_items(dnp_doc.cell(dnp_row,5))]
    dnp_d_list = dnp_d_list + [extract_items(dnp_doc.cell(dnp_row,0))]
#end for

comp_dnp_list = []

# iternate through the part numbers in the total list
for index in range(0,len(dnp_pn_list)):
    if dnp_pn_list[index] not in bom_pn_list:
        # if the part number is not present in the list of only placed
        # then move all such parts to the DNP list
        comp_dnp_list = comp_dnp_list + [dnp_d_list[index]]
        
        # empty the placed list
        dnp_d_list[index] = []
        
    else:
        # the part number is present in the placed list so check every designator
        designator_list = []
        
        # find the designator list in the other BOM
        bom_index = bom_pn_list.index(dnp_pn_list[index])
        
        # check each designator
        for designator in dnp_d_list[index]:
            # compare to the other list
            if designator not in bom_d_list[bom_index]:
                # it is not found so add the componant to the dnp list
                designator_list = designator_list + [designator]
                
                # remove the designator from this list
                dnp_d_list[index].remove(designator)
            # end if
        # end for
        
        # add to the dnp list
        comp_dnp_list = comp_dnp_list + [designator_list]
        
    # end if 
# end for

starting_dir = 'C:\Users\Asteria\Dropbox\Satellite\Pumpkin PCBs\ADCS Interface Module 2 (01845B)'

assy_filename = ''

for filename in os.listdir(starting_dir):
    if ('ASSY' in filename) and ('REV' in filename):
        assy_filename = filename
    # end if
# end for
    

assy_doc = openpyxl.load_workbook(starting_dir + '\\' + assy_filename)

bom_sheet = assy_doc.get_sheet_by_name('BOM')

for i in range(0,dnp_doc.nrows):
    for j in range(0,dnp_doc.ncols):
        if (i > 5):
            # into BOM information
            if (j == 0) and (i < dnp_doc.nrows):
                # add the designators to place
                bom_sheet.cell(i+1,j+1).value = ', '.join(dnp_d_list[i-6])
                
            elif ((j == 1) and (i < dnp_doc.nrows-1)):
                # add the designators not to place
                bom_sheet.cell(i+1,j+1).value = ', '.join(comp_dnp_list[i-6])   
                
                
            else:
                # copy the BOM info from the BOM
                bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)
            # end if
            
            bom_sheet.cell(i+1,j+1).border = thin_border
        
        else:
            # copy the BOM info from the BOM
            bom_sheet.cell(i+1,j+1).value = dnp_doc.cell_value(i,j)            
        # end if
    # end for
# end for

assy_doc.save(starting_dir + '\\test.xlsx')
                


                
        
        
 